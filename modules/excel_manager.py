import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
RESULTS_DIR = BASE_DIR / "CLASS"   # FINAL STORAGE ROOT


# -----------------------------------
# Ensure folder exists
# -----------------------------------
def ensure_folder(path: Path):
    path.mkdir(parents=True, exist_ok=True)
    return path


# -----------------------------------
# Normalize subject folder name
# -----------------------------------
def normalize_subject(subject: str):
    if not subject:
        return "Unknown"
    return subject.strip().replace(" ", "_").replace("-", "_").capitalize()


# -----------------------------------
# Build Excel file path
# CLASS/SS1/Biology/results.xlsx
# -----------------------------------
def get_excel_path(class_category: str, subject: str):
    class_folder = ensure_folder(RESULTS_DIR / class_category.upper())
    subject_folder = ensure_folder(class_folder / normalize_subject(subject))
    excel_file = subject_folder / "results.xlsx"
    return excel_file


# -----------------------------------
# Auto-repair headers — SAFE VERSION
# -----------------------------------
EXPECTED_HEADERS = [
    "Student Name", "Admission No", "Class", "Subject",
    "Score (%)", "Correct", "Total", "Status", "Submitted At"
]


def repair_missing_headers(ws):
    """
    Safely handle empty sheets or missing headers.
    Returns the cleaned header row or empty list.
    """

    rows = list(ws.iter_rows(values_only=True))

    # EMPTY SHEET — add headers and return them
    if not rows:
        for c, val in enumerate(EXPECTED_HEADERS, start=1):
            ws.cell(row=1, column=c).value = val
        return EXPECTED_HEADERS

    first_row = list(rows[0])

    # If row is empty: add headers
    if all(cell is None for cell in first_row):
        for c, val in enumerate(EXPECTED_HEADERS, start=1):
            ws.cell(row=1, column=c).value = val
        return EXPECTED_HEADERS

    # If existing headers are numeric → broken → fix
    if all(isinstance(col, int) for col in first_row):
        for c, val in enumerate(EXPECTED_HEADERS, start=1):
            ws.cell(row=1, column=c).value = val
        return EXPECTED_HEADERS

    # Otherwise return valid header
    return first_row


# -----------------------------------
# Append result to Excel (always safe)
# -----------------------------------
def append_result_to_excel(result: dict):

    class_cat = result.get("class_category", "UNKNOWN").upper()
    subject = result.get("subject", "UNKNOWN")
    excel_path = get_excel_path(class_cat, subject)

    # Create workbook if missing
    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(EXPECTED_HEADERS)
        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    repair_missing_headers(ws)

    score_percent = result.get("score", 0)
    status = "PASS" if score_percent >= 50 else "FAIL"

    ws.append([
        result.get("full_name"),
        result.get("admission_number"),
        result.get("class_name"),
        subject,
        f"{score_percent}%",
        result.get("correct", 0),
        result.get("total", 0),
        status,
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])

    wb.save(excel_path)
    return True


# -----------------------------------
# Read Excel results safely
# -----------------------------------
def read_results(class_category: str, subject: str):

    excel_path = get_excel_path(class_category, subject)

    if not excel_path.exists():
        return []

    wb = load_workbook(excel_path)
    ws = wb.active

    headers = repair_missing_headers(ws)

    rows = list(ws.iter_rows(values_only=True))

    # Still empty after repair → return []
    if len(rows) < 2:
        return []

    # If wrong header count → enforce expected
    if len(headers) != len(EXPECTED_HEADERS):
        headers = EXPECTED_HEADERS

    data = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        data.append(d)

    return data


# -----------------------------------
# Student history lookup
# -----------------------------------
def read_student_history(full_name: str, class_category: str):
    results = []
    class_dir = RESULTS_DIR / class_category.upper()

    if not class_dir.exists():
        return results

    for subject_folder in class_dir.iterdir():
        excel_file = subject_folder / "results.xlsx"

        if excel_file.exists():
            wb = load_workbook(excel_file)
            ws = wb.active

            headers = repair_missing_headers(ws)
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                continue

            for row in rows[1:]:
                row_dict = dict(zip(headers, row))
                if str(row_dict.get("Student Name", "")).strip().upper() == full_name.strip().upper():
                    results.append(row_dict)

    return results
